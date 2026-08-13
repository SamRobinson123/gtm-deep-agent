"""Scrape the "Products by Geo" tab out of the weekly GTM Exec Pipe Balance
Summaries workbooks into one tidy long-format CSV.

Source: the synced SharePoint library
    <ONEDRIVE>/GTM Exec Package/FY<yy> Q<q>/Week <n>/Pipe Balance Summaries*.xlsx

The "Products by Geo" tab only exists from FY25 Q2 (Week 15, ~May 2025) onward,
so that is the scrape window. Each weekly workbook lays the tab out as several
quarter-blocks side by side (Total Pipe | LS Pipe | ACV | ACV Target | coverage…)
with rows = product, each followed by its AMS / EMEA / APAC (sometimes EMEA/APAC)
breakdown. Older files in the window shift the whole block one column left, so
the row-label columns are derived relative to the first detected block.

Many of the older workbooks are OneDrive "online-only" placeholders; copying them
to a temp path force-hydrates the bytes so openpyxl can read them.

Output: data/inputs/gtm_exec_products_by_geo.csv
    quarter, fy, q, source_quarter, source_week, week_in_quarter, is_current_quarter,
    product_raw, product, geo, total_pipe, ls_pipe, acv, acv_target, source_file
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

# ── locate the GTM Exec Package next to the OneDrive-synced repo ───────────────
_REPO = Path(__file__).resolve().parents[1]
# repo lives at .../Strategic Analytics - Strategic Analytics/Sam Robinson/Coverage Curve Analysis
_STRATEGIC_ROOT = _REPO.parents[1]
GTM_ROOT = _STRATEGIC_ROOT / "GTM Exec Package"

OUT_CSV = _REPO / "data" / "inputs" / "gtm_exec_products_by_geo.csv"

# scrape window: the tab first appears in FY25 Q2
QUARTERS = [
    ("FY25", "Q2"), ("FY25", "Q3"), ("FY25", "Q4"),
    ("FY26", "Q1"), ("FY26", "Q2"),
]

GEO_TOKENS = {
    "AMS", "EMEA", "APAC", "EMEA/APAC", "EMEA/AP", "APAC/EMEA",
    "PubSec", "Public Sector", "LATAM",
}

# tab product name → canonical snapshot product (see data/inputs/loaders.py)
PRODUCT_CANONICAL = {
    "tosca": "Tosca",
    "testim": "Testim",
    "tricentis di": "Data Integrity",
    "tosca di": "Data Integrity",
    "di": "Data Integrity",
    "data integrity": "Data Integrity",
    "livecompare": "LiveCompare",
    "neoload": "NeoLoad",
    "qtest": "qTest",
    "sealights": "Sealights",
    "vera": "Vera",
    "agnostic service": "Recurring Services",
    "agnostic services": "Recurring Services",
    "no_product_assigned": "Other",
    "no product assigned": "Other",
    "services/unassigned": "Other",
    "other/unassigned": "Other",
}

_VERSION_RE = re.compile(r"\bv(\d+)\b", re.I)
_BAD_TOKENS = ("do not use", "old", "copy", "temp", "archive", " - bjs", "prelim")


def _norm_quarter(s) -> str | None:
    """'Q4'25' / 'Q4-2025' / 'FY25/ Q4' → 'FY25 Q4'."""
    if s is None:
        return None
    s = str(s).strip()
    m = re.search(r"Q([1-4]).{0,3}'\s*(\d{2})", s)          # Q4'25
    if m:
        return f"FY{m.group(2)} Q{m.group(1)}"
    m = re.search(r"Q([1-4])[-/ ]?20(\d{2})", s)            # Q4-2025
    if m:
        return f"FY{m.group(2)} Q{m.group(1)}"
    m = re.search(r"FY(\d{2}).{0,4}Q([1-4])", s)            # FY25/ Q4
    if m:
        return f"FY{m.group(1)} Q{m.group(2)}"
    return None


def _num(v):
    """Coerce a cell to float, mapping Excel error strings (#N/A, #REF!) to NaN."""
    if v is None:
        return float("nan")
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("#") or s == "":
        return float("nan")
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return float("nan")


def _hydrate_open(path: Path):
    """Open a workbook, force-hydrating OneDrive online-only placeholders."""
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        tmp = Path(tempfile.gettempdir()) / "pipebal_hydrate.xlsx"
        shutil.copyfile(path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True, read_only=True)


def _geo_tab(wb):
    for n in wb.sheetnames:
        nl = n.lower()
        if "product" in nl and "geo" in nl:
            return wb[n]
    return None


def _rank_files(week_dir: Path) -> list[Path]:
    """Rank the Pipe Balance workbooks in a week folder, best first.

    Prefers files without a 'bad' token (DO NOT USE / old / Copy / …), then the
    highest explicit version, then the most recently modified — so the latest
    refresh wins. Returns a list so the caller can fall back to the next
    candidate when the top pick is locked open in Excel or fails to open.
    """
    cands = [p for p in week_dir.glob("*.xlsx")
             if "pipe balance" in p.name.lower()]

    def score(p: Path):
        name = p.name.lower()
        bad = any(t in name for t in _BAD_TOKENS)
        vm = _VERSION_RE.search(name)
        version = int(vm.group(1)) if vm else 0
        return (not bad, version, p.stat().st_mtime)

    return sorted(cands, key=score, reverse=True)


def _isnan(v) -> bool:
    return v is None or (isinstance(v, float) and v != v)


# emitted under product_raw for the sheet's own grand-total row (used only for
# validation — dropped before the CSV is written)
SHEET_TOTAL = "__SHEET_TOTAL__"


def _parse_tab(ws):
    """Yield dict rows from a Products-by-Geo worksheet.

    Two passes: first classify each label row once (product-total / geo /
    single-row product / sheet-total) off the left label columns; then read each
    quarter-block's 4 metric columns at those rows. A product's "All" total is
    the sheet's own subtotal cell when numeric, else the **sum of its geo rows**
    — the sheet leaves some subtotal cells as #N/A (e.g. Agnostic Services) even
    though the geo rows carry real pipe, which otherwise under-counts the total.
    Single-row products with no geo split (No_Product_Assigned / Services-
    Unassigned) are kept as their own "All" row. The sheet's "Total" row is
    emitted under product_raw=SHEET_TOTAL for reconciliation.
    """
    grid = [[c.value for c in row]
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60),
                                    max_col=min(ws.max_column, 60))]

    def cell(r, c):
        return grid[r - 1][c - 1] if 0 <= r - 1 < len(grid) and 0 <= c - 1 < len(grid[r - 1]) else None

    # blocks: each column in row 4 labelled "Total Pipe" starts a quarter block
    blocks = []
    for c in range(1, len(grid[3]) + 1 if len(grid) >= 4 else 1):
        if str(cell(4, c)).strip().lower() == "total pipe":
            q = _norm_quarter(cell(3, c)) or _norm_quarter(cell(1, c))
            blocks.append((c, q))
    if not blocks:
        return []

    # Label column sits just left of the first data block (col D in the older
    # FY25 Q2/Q3 layout, col E once a blank spacer column appears). The product
    # name (geo sub-rows) is always col B and the product[+geo] key always col C
    # in BOTH layouts — only the label column shifts — so pin those.
    first_col = blocks[0][0]
    label_col = first_col - 1          # geo / product-total / 'Total' label
    prod_b_col = 2                     # product name on geo sub-rows (col B)
    prod_c_col = 3                     # product[+geo] key (col C)

    # ── pass 1: classify label rows → (row, product_raw, geo, is_total)
    classified = []
    for r in range(5, len(grid) + 1):
        b = cell(r, prod_b_col); c_ = cell(r, prod_c_col); e = cell(r, label_col)
        bs = str(b).strip() if b is not None else ""
        cs = str(c_).strip() if c_ is not None else ""
        es = str(e).strip() if e is not None else ""
        if es.lower() == "total" or cs.lower() == "total":
            classified.append((r, SHEET_TOTAL, "All", True))
        elif es in GEO_TOKENS and bs:
            classified.append((r, bs, es, False))               # geo sub-row
        elif bs and cs and bs == cs and es not in GEO_TOKENS:
            classified.append((r, bs, "All", False))            # single-row product
        elif not bs and cs and es == cs:
            classified.append((r, cs, "All", False))            # product-total row

    # ── pass 2: per block, per product → geo rows + a geo-sum-backed "All"
    rows = []
    for c, q in blocks:
        if q is None:
            continue
        metrics = lambda r: (_num(cell(r, c)), _num(cell(r, c + 1)),
                             _num(cell(r, c + 2)), _num(cell(r, c + 3)))
        geo_rows = {}   # product -> [(geo, metrics)]
        all_cell = {}   # product -> metrics from its own subtotal/single row
        for r, prod, geo, is_total in classified:
            m = metrics(r)
            if is_total:
                rows.append(dict(block_quarter=q, product_raw=SHEET_TOTAL, geo="All",
                                 total_pipe=m[0], ls_pipe=m[1], acv=m[2], acv_target=m[3]))
            elif geo == "All":
                all_cell[prod] = m
                geo_rows.setdefault(prod, [])
            else:
                geo_rows.setdefault(prod, []).append((geo, m))
                rows.append(dict(block_quarter=q, product_raw=prod, geo=geo,
                                 total_pipe=m[0], ls_pipe=m[1], acv=m[2], acv_target=m[3]))

        for prod, geos in geo_rows.items():
            am = all_cell.get(prod)

            def total(i, am=am, geos=geos):
                if am is not None and not _isnan(am[i]):
                    return am[i]
                vals = [g[1][i] for g in geos if not _isnan(g[1][i])]
                return sum(vals) if vals else float("nan")

            rows.append(dict(block_quarter=q, product_raw=prod, geo="All",
                             total_pipe=total(0), ls_pipe=total(1),
                             acv=total(2), acv_target=total(3)))
    return rows


def scrape() -> pd.DataFrame:
    if not GTM_ROOT.exists():
        raise FileNotFoundError(f"GTM Exec Package not found at {GTM_ROOT}")

    records = []
    for fy, q in QUARTERS:
        qdir = GTM_ROOT / f"{fy} {q}"
        if not qdir.exists():
            print(f"  ! missing quarter folder: {qdir.name}")
            continue
        source_quarter = f"{fy} {q}"
        q_start_week = {"Q1": 1, "Q2": 14, "Q3": 27, "Q4": 40}[q]

        for week_dir in sorted(qdir.glob("Week *"),
                               key=lambda p: int(re.search(r"(\d+)", p.name).group(1))):
            wm = re.search(r"(\d+)", week_dir.name)
            folder_week = int(wm.group(1)) if wm else None
            wb = f = None
            for cand in _rank_files(week_dir):
                try:
                    wb = _hydrate_open(cand)
                    f = cand
                    break
                except Exception as e:
                    print(f"  ! cannot open {cand.name}: {e!r} — trying next candidate")
            if wb is None:
                continue
            ws = _geo_tab(wb)
            if ws is None:
                continue
            parsed = _parse_tab(ws)
            if not parsed:
                print(f"  ! no rows parsed: {source_quarter} / {week_dir.name}")
                continue

            # week-in-quarter from the FOLDER's absolute FY week (the filename's
            # "Beginning of wk N" text is unreliable in FY26 — e.g. a Week 22
            # folder labelled "...Beginning of wk 21"). Folder Week N is the
            # beginning-of-week-N snapshot, matching the dashboard's
            # beginning-of-week pinning. Normalise absolute→quarter-relative and
            # clip to 1-13 (FY26 Q1 has a Week 14 wrap-up that pins to wk 13).
            if folder_week is None:
                wiq = None
            else:
                wiq = min(max(folder_week - q_start_week + 1, 1), 13)

            for rec in parsed:
                rec.update({
                    "source_quarter": source_quarter,
                    "source_week": folder_week,
                    "week_in_quarter": wiq,
                    "is_current_quarter": rec["block_quarter"] == source_quarter,
                    "source_file": f.name,
                })
                records.append(rec)

    df = pd.DataFrame.from_records(records)
    if df.empty:
        return df

    # ── reconcile: per current-quarter week, sum of product "All" rows vs the
    # sheet's own "Total" row. A gap means a product row is being missed.
    rec_v = df[df["is_current_quarter"] & (df["geo"] == "All")]
    prod_sum = (rec_v[rec_v["product_raw"] != SHEET_TOTAL]
                .groupby(["source_quarter", "source_week"])["total_pipe"].sum())
    sheet_tot = (rec_v[rec_v["product_raw"] == SHEET_TOTAL]
                 .groupby(["source_quarter", "source_week"])["total_pipe"].first())
    rec_cmp = pd.concat([prod_sum.rename("prod"), sheet_tot.rename("sheet")], axis=1)
    rec_cmp["diff"] = (rec_cmp["prod"] - rec_cmp["sheet"]).abs()
    bad = rec_cmp[rec_cmp["diff"] > 0.5]
    n_ok = int((rec_cmp["diff"] <= 0.5).sum())
    print(f"  reconcile vs sheet Total row: {n_ok}/{rec_cmp['sheet'].notna().sum()} "
          f"current-quarter weeks tie (open pipe, ±0.5M)")
    if not bad.empty:
        print("  ! weeks NOT tying to the sheet Total row:")
        print(bad.round(2).to_string())

    # Keep the sheet "Total" rows as product='__TOTAL__' (the file's bold Total
    # line — used as the product page's grand total, which Sam wants to tie to
    # exactly). Product rows are filtered to recognised products; report anything
    # else dropped so a genuinely new product is never silently lost
    # (junk: CHECK / Q4 / geo tokens / numbers).
    is_total = df["product_raw"] == SHEET_TOTAL
    keep = is_total | df["product_raw"].astype(str).str.strip().str.lower().isin(PRODUCT_CANONICAL)
    dropped = sorted(df.loc[~keep, "product_raw"].astype(str).unique().tolist())
    if dropped:
        print(f"  dropped non-product rows: {dropped}")
    df = df[keep].copy()
    df["product"] = df["product_raw"].astype(str).str.strip().str.lower().map(PRODUCT_CANONICAL)
    df.loc[df["product_raw"] == SHEET_TOTAL, "product"] = "__TOTAL__"
    df = df.rename(columns={"block_quarter": "quarter"})

    # ── Cross-quarter endpoints (Sam, 2026-06-11) ─────────────────────────────
    # A quarter's wk1 (start-of-quarter) snapshot and its wk13 (end) live in the
    # ADJACENT quarter's weekly files: a quarter's own folders run wk2..wk12 (its
    # wk1 = the prior quarter's last file's forward block; its wk13 = the next
    # quarter's first file's trailing block). Promote those so the curves are
    # complete and tie to the files (= what the GTM decks show).
    def _promote_endpoint(this_q, other_q, week, pick):
        m = (df["quarter"] == this_q) & (df["source_quarter"] == other_q)
        if not m.any():
            return
        sw = df.loc[m, "source_week"].max() if pick == "last" else df.loc[m, "source_week"].min()
        sel = m & (df["source_week"] == sw)
        df.loc[sel, "is_current_quarter"] = True
        df.loc[sel, "week_in_quarter"] = week

    # FY25 Q4: own folders are labelled one short (wk1-11 → true wk2-12); shift,
    # then pull wk1 (from FY25 Q3's last file) and wk13 (from FY26 Q1's first file
    # — total 31.1 open / 25.6 ACV, matching the Week 51 deck's product slide).
    q4 = "FY25 Q4"
    cur4 = (df["quarter"] == q4) & df["is_current_quarter"] & df["week_in_quarter"].notna()
    df.loc[cur4, "week_in_quarter"] = df.loc[cur4, "week_in_quarter"] + 1
    _promote_endpoint(q4, "FY25 Q3", 1, "last")
    _promote_endpoint(q4, "FY26 Q1", 13, "first")

    # FY26 Q2 (in-flight): own folders are labelled one HIGH — the "Week 15"
    # folder holds the Beginning-of-wk-14 snapshot, i.e. Q2 wk1 (start-of-quarter,
    # $103.2M open), not wk2. (Q2 starts at FY wk14; "Beginning of wk N" in the
    # filename = Q2 wk N-13.) The generic absolute-week formula over-counts by 1,
    # so shift current weeks down: Week 15→wk1 … Week 25→wk11. Do NOT promote wk1
    # from FY26 Q1's last file — that block is the Beginning-of-wk-13 (a Q1 week)
    # forward projection of Q2 ($105.3M), not the actual Q2 start.
    q2 = "FY26 Q2"
    cur2 = (df["quarter"] == q2) & df["is_current_quarter"] & df["week_in_quarter"].notna()
    df.loc[cur2, "week_in_quarter"] = df.loc[cur2, "week_in_quarter"] - 1

    df["fy"] = df["quarter"].str.slice(0, 4)
    df["q"] = df["quarter"].str.slice(5, 7)
    return df[[
        "quarter", "fy", "q", "source_quarter", "source_week", "week_in_quarter",
        "is_current_quarter", "product_raw", "product", "geo",
        "total_pipe", "ls_pipe", "acv", "acv_target", "source_file",
    ]]


def main():
    df = scrape()
    if df.empty:
        print("No data scraped.")
        return
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_CSV, index=False)
    print(f"\nWrote {len(df):,} rows -> {OUT_CSV}")
    print(f"weeks: {df.groupby('source_quarter')['source_week'].nunique().to_dict()}")
    cur = df[df["is_current_quarter"] & (df["geo"] == "All")]
    print(f"\ncurrent-quarter product-total rows: {len(cur)}")
    print("products seen:", sorted(df["product_raw"].unique().tolist()))
    print("geos seen:", sorted(df["geo"].unique().tolist()))


if __name__ == "__main__":
    main()

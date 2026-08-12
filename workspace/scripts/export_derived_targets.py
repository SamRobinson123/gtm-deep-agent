import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

RUN_DIR = r"C:\Dev V2\gtm-deep-agent\workspace\runs\2026-08-12T000723Z_e160d4"
OUT_PATH = r"C:\Dev V2\gtm-deep-agent\workspace\exports\pipe_create_derived_targets_Q3_Q4_FY26_2026-08-11_territory.xlsx"

df = pd.read_csv(f"{RUN_DIR}\\derived_pipe_create.csv")

summary = pd.DataFrame([
    {"Quarter": "Q3 FY26", "Derived target": 449844802.0, "Published target": 201789917.58,
     "Delta $": 449844802.0 - 201789917.58, "Delta %": (449844802.0 - 201789917.58) / 201789917.58},
    {"Quarter": "Q4 FY26", "Derived target": 537257476.0, "Published target": 192223413.39,
     "Delta $": 537257476.0 - 192223413.39, "Delta %": (537257476.0 - 192223413.39) / 192223413.39},
])

floor_summary = pd.DataFrame([{
    "Floor-driven $": 28592284.33,
    "Floor-driven %": 0.029,
    "Rows floor-bound": 8,
    "Rows gap-bound": 44,
    "Note": "Floor-driven = territory may not create less than the same quarter last year. Gap-driven = the bookings target requires it.",
}])

assumptions = pd.DataFrame([
    {"Assumption": "Sales-cycle / win-rate fitting window", "Value": "2024-07-01 to 2026-06-30 (trailing 8 completed quarters)"},
    {"Assumption": "Slip measurement quarters", "Value": "Q3 FY25, Q4 FY25, Q1 FY26, Q2 FY26"},
    {"Assumption": "Grain", "Value": "Territory"},
    {"Assumption": "Status", "Value": "Not a documented default -- waterfall doc open question 3 leaves the fitting window unestablished. Chosen for this run; not the published target."},
    {"Assumption": "Source data", "Value": "Cached data/sku_nacv.parquet, pulled 2026-08-11 05:10Z"},
    {"Assumption": "Run ID", "Value": "2026-08-12T000723Z_e160d4"},
])

dollar_cols = [
    "bookings_target", "closed_won", "expected_from_existing_pipe",
    "sales_cycle_tail_from_earlier_quarters", "gap", "required_by_gap",
    "historic_floor", "pipe_create_target",
]
pct_cols = ["yield_per_dollar", "in_quarter_win_rate", "pre_q_win_rate", "q0_weight"]

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
    floor_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=len(summary) + 3)
    assumptions.to_excel(writer, sheet_name="Assumptions", index=False)
    df.to_excel(writer, sheet_name="Territory Detail", index=False)

    wb = writer.book

    ws = wb["Summary"]
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=len(summary) + 4, max_row=len(summary) + 4):
        for cell in row:
            cell.font = Font(bold=True)
    for col_name in ["Derived target", "Published target", "Delta $"]:
        col_idx = summary.columns.get_loc(col_name) + 1
        for r in range(2, len(summary) + 2):
            ws.cell(row=r, column=col_idx).number_format = "#,##0"
    delta_pct_idx = summary.columns.get_loc("Delta %") + 1
    for r in range(2, len(summary) + 2):
        ws.cell(row=r, column=delta_pct_idx).number_format = "0.0%"
    floor_row = len(summary) + 5
    ws.cell(row=floor_row, column=1).number_format = "#,##0"
    ws.cell(row=floor_row, column=2).number_format = "0.0%"
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, length + 2)

    ws2 = wb["Assumptions"]
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 90

    ws3 = wb["Territory Detail"]
    for row in ws3.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    ws3.freeze_panes = "A2"
    for col_name in dollar_cols:
        col_idx = df.columns.get_loc(col_name) + 1
        for r in range(2, len(df) + 2):
            ws3.cell(row=r, column=col_idx).number_format = "#,##0"
    for col_name in pct_cols:
        col_idx = df.columns.get_loc(col_name) + 1
        for r in range(2, len(df) + 2):
            ws3.cell(row=r, column=col_idx).number_format = "0.0%"
    for col_cells in ws3.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:1] + col_cells[1:50])
        ws3.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(10, length + 2), 40)

print("wrote", OUT_PATH)
print("rows:", len(df))

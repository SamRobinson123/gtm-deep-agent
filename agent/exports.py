"""File exports — the ONLY place this agent writes deliverables.

DESIGN RULE, load-bearing: every path is resolved through `export_path()`, which
confines writes to `workspace/exports/`. The agent chooses a *name*, never a
path. A tool that took a path would let a prompt-injected filename write into
`data/` (read-only), `<REPO>/output/` (needs explicit approval) or anywhere on
disk — so the boundary is enforced here rather than trusted to the caller,
exactly as `pipeline/queries.py` is the boundary for SQL.

Conventions are from root CLAUDE.md and are not stylistic preferences:
  - dollars `#,##0`, attainment/rates `0.0%`
  - header row frozen, columns auto-width
  - charts 150 dpi, one chart per file, titled with quarter and grain
  - NEVER overwrite an existing export — collisions get a date, then a counter
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

from pipeline import config

EXPORTS = config.ROOT / "workspace" / "exports"

# Columns whose values are dollars, and those that are rates. Matched by name so
# a new column inherits the right format without another edit here.
MONEY = re.compile(
    r"target|won|pipe|gap|floor|required|value|create|bookings|inflow|tail|"
    r"existing|amount|nacv|iacv|slip(?!_rate)", re.I)
RATE = re.compile(r"rate|weight|yield|pct|share|attainment", re.I)

_SAFE = re.compile(r"[^A-Za-z0-9._-]+")


def safe_stem(name: str, default: str = "export") -> str:
    """A filename stem with every path-ish character removed.

    Not a sanitiser that tries to detect traversal — a whitelist. `../../x`,
    `C:\\Windows\\y` and `a/b` all collapse to a flat token, so the result cannot
    escape EXPORTS no matter what produced it.
    """
    stem = _SAFE.sub("_", str(name or "")).strip("._-")
    return (stem or default)[:120]


def export_path(name: str, suffix: str) -> Path:
    """A never-overwriting path inside workspace/exports/.

    Collision policy from CLAUDE.md: suffix with the date, and if THAT exists
    too, add a counter. Silently overwriting a deliverable someone has already
    opened or sent on is the failure this prevents.
    """
    EXPORTS.mkdir(parents=True, exist_ok=True)
    stem = safe_stem(name)
    p = EXPORTS / f"{stem}{suffix}"
    if not p.exists():
        return _confine(p)
    dated = EXPORTS / f"{stem}_{pd.Timestamp.today():%Y-%m-%d}{suffix}"
    if not dated.exists():
        return _confine(dated)
    for i in range(2, 100):
        c = EXPORTS / f"{stem}_{pd.Timestamp.today():%Y-%m-%d}_{i}{suffix}"
        if not c.exists():
            return _confine(c)
    raise RuntimeError(f"cannot find a free filename for {stem!r} in {EXPORTS}")


def _confine(p: Path) -> Path:
    """Belt and braces: prove the resolved path is under EXPORTS before returning.

    safe_stem() should make this unreachable. It is here because "should" is not
    a guarantee, and the cost of being wrong is a write outside the sandbox.
    """
    rp, root = p.resolve(), EXPORTS.resolve()
    if not str(rp).startswith(str(root) + "\\") and not str(rp).startswith(str(root) + "/"):
        raise ValueError(f"refusing to write outside {root}: {rp}")
    return rp


def _fmt_for(col: str) -> str | None:
    if RATE.search(col):
        return "0.0%"
    if MONEY.search(col):
        return "#,##0"
    return None


def write_sheet(writer, df: pd.DataFrame, sheet: str):
    """One formatted sheet: frozen header, auto-width, money and rate formats."""
    sheet = sheet[:31] or "Sheet1"          # Excel's hard limit
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    ws.freeze_panes = "A2"

    from openpyxl.styles import Alignment, Font
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for i, col in enumerate(df.columns, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        fmt = _fmt_for(str(col))
        if fmt:
            for r in range(2, len(df) + 2):
                ws.cell(row=r, column=i).number_format = fmt
        # Width from the widest rendered value, not the raw repr: a float column
        # formatted #,##0 displays far shorter than str(1234.5678901).
        body = df[col]

        def rendered(v):
            # Always a str. pandas 3's astype(str) yields the `str` dtype, which
            # keeps NaN as MISSING rather than the text "nan" — so iterating an
            # astype(str) column can still hand back a float and len() blows up.
            if pd.isna(v):
                return ""
            if fmt == "#,##0":
                return f"{v:,.0f}"
            if fmt == "0.0%":
                return f"{v:.1%}"
            return str(v)

        shown = [rendered(v) for v in body.head(500)]
        width = max([len(str(col))] + [len(x) for x in shown]) + 2
        ws.column_dimensions[letter].width = min(max(width, 9), 46)
    return ws
